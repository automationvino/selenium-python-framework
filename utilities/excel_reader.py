from openpyxl import load_workbook

def get_test_data(file_path, sheet_name):
    from openpyxl import load_workbook

    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if None in row:
            continue   # skip empty rows
        data.append(row)

    return data